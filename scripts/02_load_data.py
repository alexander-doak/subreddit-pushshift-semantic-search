"""Load subreddit data from XLSX into the reddit_vectors database."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config
import database
import openpyxl
from collections import Counter


XLSX_PATH = Path(r"F:\Emtherical\data\reddit\vector_search\subreddit_descriptions.xlsx")

REQUIRED_COLUMNS = [
    "subreddit_name",
    "url_checked",
    "final_url",
    "status_code",
    "page_title",
    "signal_banned",
    "signal_private",
    "signal_quarantined",
    "signal_over18",
    "signal_not_found",
    "description",
    "comments_size_mb",
    "submissions_size_mb",
]


def parse_bool(value):
    """Convert TRUE/FALSE strings to Python bool."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().upper() == "TRUE"


def parse_float(value):
    """Convert to float, returning None for blanks."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_int(value):
    """Convert to int, returning None for blanks."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def clean_description(value):
    """Return None for SKIP or blank descriptions, otherwise stripped text."""
    if value is None:
        return None
    text = str(value).strip()
    if text.upper() == config.SKIP_SENTINEL or text == "":
        return None
    return text


def load_xlsx(path):
    """Read the XLSX file and return a list of row dicts."""
    print(f"Opening {path} ...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        val = cell.value
        headers.append(str(val).strip().lower() if val is not None else "")

    # Check required columns
    missing = [c for c in REQUIRED_COLUMNS if c.lower() not in headers]
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        print(f"Found columns: {headers}")
        sys.exit(1)

    # Map header names to column indices
    col_idx = {name: i for i, name in enumerate(headers)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def get(col):
            return row[col_idx[col]] if col_idx[col] < len(row) else None

        rows.append({
            "subreddit_name":      str(get("subreddit_name") or "").strip(),
            "url_checked":         str(get("url_checked") or "").strip() or None,
            "final_url":           str(get("final_url") or "").strip() or None,
            "status_code":         parse_int(get("status_code")),
            "page_title":          str(get("page_title") or "").strip() or None,
            "signal_banned":       parse_bool(get("signal_banned")),
            "signal_private":      parse_bool(get("signal_private")),
            "signal_quarantined":  parse_bool(get("signal_quarantined")),
            "signal_over18":       parse_bool(get("signal_over18")),
            "signal_not_found":    parse_bool(get("signal_not_found")),
            "description":         clean_description(get("description")),
            "comments_size_mb":    parse_float(get("comments_size_mb")),
            "submissions_size_mb": parse_float(get("submissions_size_mb")),
        })

    wb.close()
    return rows


def check_duplicates(rows):
    """Check for duplicate subreddit_name values. Returns a dict of name -> count for duplicates."""
    counts = Counter(r["subreddit_name"] for r in rows)
    return {name: count for name, count in counts.items() if count > 1}


def deduplicate_rows(rows):
    """Keep only the last occurrence of each subreddit_name."""
    seen = {}
    for i, row in enumerate(rows):
        seen[row["subreddit_name"]] = i

    deduped = [rows[i] for i in sorted(seen.values())]
    return deduped


def insert_rows(rows):
    """Insert rows into the subreddits table."""
    insert_sql = """
        INSERT INTO subreddits (
            subreddit_name, url_checked, final_url, status_code, page_title,
            signal_banned, signal_private, signal_quarantined,
            signal_over18, signal_not_found,
            description, comments_size_mb, submissions_size_mb
        ) VALUES (
            %(subreddit_name)s, %(url_checked)s, %(final_url)s, %(status_code)s, %(page_title)s,
            %(signal_banned)s, %(signal_private)s, %(signal_quarantined)s,
            %(signal_over18)s, %(signal_not_found)s,
            %(description)s, %(comments_size_mb)s, %(submissions_size_mb)s
        )
    """

    batch_size = 500
    total = len(rows)

    with database.get_cursor() as cur:
        for i in range(0, total, batch_size):
            batch = rows[i : i + batch_size]
            cur.executemany(insert_sql, batch)
            loaded = min(i + batch_size, total)
            print(f"  Inserted {loaded:,} / {total:,} rows ...")

    print(f"Done. {total:,} rows inserted.")


def main():
    # Pre-flight check
    print("=" * 60)
    print("SUBREDDIT DATA LOADER")
    print("=" * 60)
    print()
    print(f"Source file: {XLSX_PATH}")
    print()
    print("Required columns:")
    for col in REQUIRED_COLUMNS:
        print(f"  - {col}")
    print()
    print("Please verify:")
    print("  1. The XLSX file exists at the path above")
    print("  2. Row 1 contains the required column headers")
    print("  3. Boolean columns use TRUE / FALSE values")
    print(f"  4. Descriptions with value '{config.SKIP_SENTINEL}' will be stored as NULL")
    print()

    input("Press ENTER to continue (or Ctrl+C to abort) ... ")
    print()

    if not XLSX_PATH.exists():
        print(f"ERROR: File not found: {XLSX_PATH}")
        sys.exit(1)

    # Check if table already has data
    with database.get_cursor() as cur:
        cur.execute("SELECT COUNT(*) AS cnt FROM subreddits;")
        existing = cur.fetchone()["cnt"]

    if existing > 0:
        print(f"WARNING: The subreddits table already has {existing:,} rows.")
        resp = input("Type 'CLEAR' to delete them first, or ENTER to append: ").strip()
        if resp == "CLEAR":
            with database.get_cursor() as cur:
                cur.execute("DELETE FROM subreddits;")
            print("Existing rows deleted.")
        print()

    rows = load_xlsx(XLSX_PATH)
    print(f"Parsed {len(rows):,} rows from XLSX.")

    desc_count = sum(1 for r in rows if r["description"] is not None)
    skip_count = len(rows) - desc_count
    print(f"  Descriptions: {desc_count:,} populated, {skip_count:,} null/skipped")
    print()

    # Check for duplicate subreddit_name values in the input
    duplicates = check_duplicates(rows)

    if duplicates:
        total_extra = sum(count - 1 for count in duplicates.values())
        print(f"WARNING: Found {len(duplicates):,} subreddit names that appear more than once "
              f"({total_extra:,} extra rows total).")
        print()

        preview = sorted(duplicates.items(), key=lambda x: x[1], reverse=True)
        show_count = min(20, len(preview))
        print(f"  Top {show_count} duplicates:")
        for name, count in preview[:show_count]:
            print(f"    r/{name} -- {count} occurrences")
        if len(preview) > show_count:
            print(f"    ... and {len(preview) - show_count} more")
        print()

        print("Options:")
        print("  D = Deduplicate (keep last occurrence of each name, drop the rest)")
        print("  A = Abort")
        print()
        resp = input("Enter choice (D/A): ").strip().upper()

        if resp == "D":
            before = len(rows)
            rows = deduplicate_rows(rows)
            after = len(rows)
            print(f"  Deduplicated: {before:,} -> {after:,} rows ({before - after:,} removed).")
            print()
        else:
            print("Aborted by user.")
            sys.exit(0)
    else:
        print("No duplicate subreddit names found in input. Good to go.")
        print()

    insert_rows(rows)

    # Final verification
    with database.get_cursor() as cur:
        cur.execute("SELECT COUNT(*) AS total FROM subreddits;")
        total = cur.fetchone()["total"]
        cur.execute("SELECT COUNT(*) AS with_desc FROM subreddits WHERE description IS NOT NULL;")
        with_desc = cur.fetchone()["with_desc"]

    print()
    print(f"Verification: {total:,} rows in table, {with_desc:,} with descriptions.")


if __name__ == "__main__":
    main()